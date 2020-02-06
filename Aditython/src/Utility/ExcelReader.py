__author__ = 'Aditya Roy'

import openpyxl
import pandas as pd
from pandas import *
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import datetime

#utility function written using openpyxl
def r_count(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row

def c_count(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column

def readData(file,sheetName,rowNum,columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum, column=columnNum).value

def writeData(file,sheetName,rowNum,columnNum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum, column=columnNum).value = data
    workbook.save(file)


#utility function written using panda
def getData(path, sheetName):
    data = read_excel(path, sheet_name=sheetName)
    return data

def updateResult(path,sheetName,columnName,passResult,rowNum,color):
    df = pd.DataFrame({columnName:[passResult]})
    wb = load_workbook(path)
    ws = wb[sheetName]

    #for cell formatting
    redFill = PatternFill(start_color='00FF0000',fill_type='solid')
    greenFill = PatternFill(start_color='0000FF00',fill_type='solid')

    for index, row in df.iterrows():
        # %d for decimal, %s for string # Cell with start C2,C3,C4 and continue)
        cell = 'C%d' % (index + 2 + rowNum)
        ws[cell] = row[columnName]
        if color=='Red':
            ws[cell].fill = redFill
        elif color=='Green':
            ws[cell].fill = greenFill
    wb.save(path)

def updateExecutionCompletion(path,sheetName,rowNum):
    df = pd.DataFrame({'ExecutionTime':[str(datetime.datetime.now())]})
    wb = load_workbook(path)
    ws = wb[sheetName]

    for index, row in df.iterrows():
        # %d for decimal, %s for string # Cell with start D2,D3,D4 and continue)
        cell = 'D%d' % (index + 2 + rowNum)
        ws[cell] = row['ExecutionTime']
    wb.save(path)
